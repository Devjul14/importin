"""
cleansing.py — Utility functions for Excel data cleansing.

All functions are pure: they accept a DataFrame and return a new DataFrame
(or supporting data), never mutating the input.
"""

from __future__ import annotations

import io
import re
from typing import Any

import pandas as pd
import openpyxl


# ─────────────────────────────────────────────
# 1. Load Excel with un-merged cells
# ─────────────────────────────────────────────

def load_excel_unmerged(
    file: Any,
    sheet_name: str | int = 0,
    header_row: int = 0,          # 0-based row index used as header
) -> pd.DataFrame:
    """
    Load an Excel sheet and forward-fill merged cells before parsing.

    openpyxl stores merged cell ranges: all cells inside the range are empty
    except the top-left. This function fills those empty cells with the value
    from the top-left before handing to pandas.
    """
    if hasattr(file, "read"):
        raw = file.read()
        file.seek(0)
    else:
        from pathlib import Path
        raw = Path(file).read_bytes()

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name] if sheet_name else wb.active

    # Un-merge: fill each merged region with the top-left value
    merged_ranges = list(ws.merged_cells.ranges)
    for merge_range in merged_ranges:
        top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in ws.iter_rows(
            min_row=merge_range.min_row, max_row=merge_range.max_row,
            min_col=merge_range.min_col, max_col=merge_range.max_col,
        ):
            for cell in row:
                cell.value = top_left_value

    # Convert to DataFrame via values
    data = list(ws.values)
    if not data:
        return pd.DataFrame()

    if header_row >= len(data):
        return pd.DataFrame()

    headers = [str(c) if c is not None else f"Col_{i}" for i, c in enumerate(data[header_row])]
    rows = data[header_row + 1:]
    df = pd.DataFrame(rows, columns=headers)
    return df


# ─────────────────────────────────────────────
# 2. Select / drop columns
# ─────────────────────────────────────────────

def select_columns(df: pd.DataFrame, keep: list[str]) -> pd.DataFrame:
    """Keep only the specified columns (in given order)."""
    valid = [c for c in keep if c in df.columns]
    return df[valid].copy()


def drop_columns(df: pd.DataFrame, drop: list[str]) -> pd.DataFrame:
    """Drop specified columns."""
    return df.drop(columns=[c for c in drop if c in df.columns])


# ─────────────────────────────────────────────
# 3. Drop empty rows / columns
# ─────────────────────────────────────────────

def drop_empty_rows(df: pd.DataFrame, how: str = "all", subset: list[str] | None = None) -> pd.DataFrame:
    """
    Remove rows where cells are empty.
    how='all'  → drop row only if ALL cells are NaN (default)
    how='any'  → drop row if ANY cell is NaN
    subset     → check only these columns
    """
    return df.dropna(how=how, subset=subset).reset_index(drop=True)


def drop_empty_columns(df: pd.DataFrame, threshold: float = 1.0) -> pd.DataFrame:
    """
    Drop columns where NaN ratio >= threshold (default=1.0 means fully empty).
    threshold=0.5 → drop columns with ≥50% missing values.
    """
    min_count = int((1 - threshold) * len(df)) + 1
    return df.dropna(axis=1, thresh=min_count)


# ─────────────────────────────────────────────
# 4. Trim whitespace & normalize strings
# ─────────────────────────────────────────────

def trim_whitespace(df: pd.DataFrame, cols: list[str] | None = None) -> pd.DataFrame:
    """Strip leading/trailing whitespace from string columns."""
    df = df.copy()
    target = cols if cols else df.select_dtypes(include="object").columns.tolist()
    for col in target:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace("nan", pd.NA)
    return df


def normalize_whitespace(df: pd.DataFrame, cols: list[str] | None = None) -> pd.DataFrame:
    """Collapse multiple spaces into one and strip."""
    df = df.copy()
    target = cols if cols else df.select_dtypes(include="object").columns.tolist()
    for col in target:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.replace(r"\s+", " ", regex=True).replace("nan", pd.NA)
    return df


def to_uppercase(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for col in cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.upper().replace("nan", pd.NA)
    return df


def to_lowercase(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for col in cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.lower().replace("nan", pd.NA)
    return df


def to_titlecase(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for col in cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.title().replace("nan", pd.NA)
    return df


# ─────────────────────────────────────────────
# 5. Rename columns
# ─────────────────────────────────────────────

def rename_columns(df: pd.DataFrame, rename_map: dict[str, str]) -> pd.DataFrame:
    """Rename columns. rename_map = {old_name: new_name}"""
    return df.rename(columns=rename_map)


# ─────────────────────────────────────────────
# 6. Deduplicate
# ─────────────────────────────────────────────

def drop_duplicates(
    df: pd.DataFrame,
    subset: list[str] | None = None,
    keep: str = "first",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Remove duplicate rows.
    Returns (clean_df, removed_df).
    """
    mask = df.duplicated(subset=subset, keep=keep)
    return df[~mask].reset_index(drop=True), df[mask].reset_index(drop=True)


# ─────────────────────────────────────────────
# 7. Type coercion
# ─────────────────────────────────────────────

def coerce_numeric(df: pd.DataFrame, cols: list[str]) -> tuple[pd.DataFrame, dict[str, int]]:
    """
    Convert columns to numeric. Non-parseable values become NaN.
    Returns (df, {col: error_count}).
    """
    df = df.copy()
    errors: dict[str, int] = {}
    for col in cols:
        if col not in df.columns:
            continue
        before_na = df[col].isna().sum()
        df[col] = pd.to_numeric(df[col], errors="coerce")
        after_na = df[col].isna().sum()
        errors[col] = int(after_na - before_na)
    return df, errors


def coerce_datetime(
    df: pd.DataFrame,
    cols: list[str],
    fmt: str | None = None,
    output_fmt: str | None = None,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """
    Parse date columns. Non-parseable become NaT → None.
    output_fmt: if set, convert to string with strftime (e.g. '%Y-%m-%d').
    Returns (df, {col: error_count}).
    """
    df = df.copy()
    errors: dict[str, int] = {}
    for col in cols:
        if col not in df.columns:
            continue
        before_na = df[col].isna().sum()
        parsed = pd.to_datetime(df[col], format=fmt, errors="coerce")
        after_na = parsed.isna().sum()
        errors[col] = int(after_na - before_na)
        if output_fmt:
            df[col] = parsed.dt.strftime(output_fmt).where(parsed.notna(), other=None)
        else:
            df[col] = parsed
    return df, errors


# ─────────────────────────────────────────────
# 8. Find & replace values
# ─────────────────────────────────────────────

def find_replace(
    df: pd.DataFrame,
    col: str,
    find: str,
    replace_with: str,
    is_regex: bool = False,
    case_sensitive: bool = True,
) -> tuple[pd.DataFrame, int]:
    """
    Find & replace in a single column. Returns (df, replaced_count).
    """
    df = df.copy()
    if col not in df.columns:
        return df, 0
    series = df[col].astype(str)
    if is_regex:
        flags = 0 if case_sensitive else re.IGNORECASE
        new_series = series.str.replace(find, replace_with, regex=True, flags=flags)
    else:
        new_series = series.str.replace(find, replace_with, regex=False, case=case_sensitive)
    count = int((new_series != series).sum())
    df[col] = new_series.replace("nan", pd.NA)
    return df, count


# ─────────────────────────────────────────────
# 9. Fill missing values
# ─────────────────────────────────────────────

def fill_missing(
    df: pd.DataFrame,
    col: str,
    method: str,         # 'value', 'ffill', 'bfill', 'mean', 'median', 'mode'
    value: Any = None,
) -> pd.DataFrame:
    df = df.copy()
    if col not in df.columns:
        return df
    if method == "value":
        df[col] = df[col].fillna(value)
    elif method == "ffill":
        df[col] = df[col].ffill()
    elif method == "bfill":
        df[col] = df[col].bfill()
    elif method == "mean":
        df[col] = df[col].fillna(pd.to_numeric(df[col], errors="coerce").mean())
    elif method == "median":
        df[col] = df[col].fillna(pd.to_numeric(df[col], errors="coerce").median())
    elif method == "mode":
        mode_val = df[col].mode()
        if not mode_val.empty:
            df[col] = df[col].fillna(mode_val.iloc[0])
    return df


# ─────────────────────────────────────────────
# 10. Split column by delimiter
# ─────────────────────────────────────────────

def split_column(
    df: pd.DataFrame,
    col: str,
    delimiter: str,
    new_col_names: list[str],
) -> pd.DataFrame:
    """Split one column into multiple columns by delimiter."""
    df = df.copy()
    if col not in df.columns:
        return df
    parts = df[col].astype(str).str.split(delimiter, expand=True)
    for i, name in enumerate(new_col_names):
        df[name] = parts[i] if i < parts.shape[1] else None
    return df


# ─────────────────────────────────────────────
# 11. Export
# ─────────────────────────────────────────────

def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def to_csv_bytes(df: pd.DataFrame, sep: str = ",") -> bytes:
    return df.to_csv(index=False, sep=sep).encode("utf-8-sig")


# ─────────────────────────────────────────────
# 12. Summary stats for report
# ─────────────────────────────────────────────

def null_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Return per-column null count and percentage."""
    counts = df.isna().sum()
    pct = (counts / len(df) * 100).round(1)
    # Use iloc-based dtype lookup to avoid issues with duplicate column names
    dtypes = [str(df.iloc[:, i].dtype) for i in range(len(df.columns))]
    return pd.DataFrame({
        "Kolom": [str(c) for c in df.columns],
        "Null Count": counts.values,
        "Null %": pct.values,
        "Dtype": dtypes,
    })


# ─────────────────────────────────────────────
# Unpivot (melt / wide → long)
# ─────────────────────────────────────────────

def strip_unit(series: pd.Series) -> pd.Series:
    """
    Hapus satuan dari nilai string numerik.
    Contoh: '150 cm' -> '150', '2743 Yrd' -> '2743', '30 Pcs' -> '30'.
    Nilai yang tidak mengandung angka akan menjadi NaN.
    """
    extracted = series.astype(str).str.extract(r'^\s*([\d.,]+)', expand=False)
    return pd.to_numeric(extracted.str.replace(',', '', regex=False), errors='coerce')


def unpivot(
    df: pd.DataFrame,
    id_cols: list[str],
    value_cols: list[str],
    var_name: str = "kode_bahan_baku",
    value_name: str = "qty",
    drop_null_values: bool = True,
    drop_dash_values: bool = True,
    strip_units: bool = False,
    sort_by_id: bool = True,
) -> pd.DataFrame:
    """
    Ubah format wide → long (unpivot).

    Parameters
    ----------
    id_cols         : kolom yang tetap (misal: ['Nama Barang'])
    value_cols      : kolom yang akan di-unpivot (misal: ['BB00022', 'BB00003', ...])
    var_name        : nama kolom baru untuk header asli (kode bahan baku)
    value_name      : nama kolom baru untuk nilai (qty)
    drop_null_values: hapus baris di mana nilai NULL setelah melt
    drop_dash_values: hapus baris di mana nilai adalah ' -', '-', atau string dash
    """
    # Work on a positional copy to handle duplicate column names safely.
    # For each selected name in value_cols, collect ALL positional occurrences
    # in the original DataFrame (multiselect only returns unique names).
    sub = df[id_cols].copy()
    col_list = df.columns.tolist()

    # Build ordered list of (internal_name, position, original_name)
    entries: list[tuple[str, int, str]] = []
    used_positions: set[int] = set()
    selected_set = list(dict.fromkeys(str(c) for c in value_cols))  # preserve order, unique

    for col_name in selected_set:
        # collect all positions for this column name
        positions = [pos for pos, c in enumerate(col_list) if str(c) == col_name]
        for i, pos in enumerate(positions):
            if pos in used_positions:
                continue
            internal = col_name if i == 0 else f"{col_name}__{i}"
            entries.append((internal, pos, col_name))
            used_positions.add(pos)

    internal_names = [e[0] for e in entries]
    original_names = [e[2] for e in entries]

    for internal, pos, _ in entries:
        sub[internal] = df.iloc[:, pos].values

    # Avoid conflict between var_name/value_name and existing columns
    all_cols = set(sub.columns.tolist())
    safe_var = var_name
    while safe_var in all_cols:
        safe_var = safe_var + "_"
    safe_val = value_name
    while safe_val in all_cols:
        safe_val = safe_val + "_"

    melted = sub.melt(
        id_vars=id_cols,
        value_vars=internal_names,
        var_name=safe_var,
        value_name=safe_val,
    ).reset_index(drop=True)

    # Map internal names back to original names in the var column
    name_map = dict(zip(internal_names, original_names))
    melted[safe_var] = melted[safe_var].map(name_map).fillna(melted[safe_var])

    # Rename safe_var / safe_val back to intended names
    melted = melted.rename(columns={safe_var: var_name, safe_val: value_name})

    if drop_null_values:
        melted = melted[melted[value_name].notna()]

    if drop_dash_values:
        # strip whitespace then drop cells that are only dashes or empty
        mask = melted[value_name].astype(str).str.strip().isin(["-", "--", "", "nan", "None"])
        melted = melted[~mask]

    if strip_units:
        melted[value_name] = strip_unit(melted[value_name])
        if drop_null_values:
            melted = melted[melted[value_name].notna()]

    if sort_by_id and id_cols:
        melted = melted.sort_values(by=id_cols, kind="stable").reset_index(drop=True)

    return melted.reset_index(drop=True)
